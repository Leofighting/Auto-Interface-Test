# -*- coding:utf-8 -*-
__author__ = "leo"

import os

import openpyxl

base_path = os.path.dirname(os.getcwd())


class HandleExcel:
    """处理 Excel 文件"""

    @staticmethod
    def load_excel():
        """
        加载 Excel 文件
        :return: Excel 文件
        """
        open_excel = openpyxl.load_workbook(base_path + "/case/case.xlsx")
        return open_excel

    def get_sheet_data(self, index=None):
        """
        获取 Excel 文件的工作表数据
        :param index: 第几个工作表：工作表索引值
        :return: 对应工作表数据
        """
        sheet_name = self.load_excel().sheetnames

        if index is None:
            index = 0

        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, col):
        """
        获取单元格内容
        :param row: 行
        :param col: 列
        :return: 单元格内容
        """
        data = self.get_sheet_data().cell(row=row, column=col).value
        return data

    def get_rows(self):
        """
        获取工作表行数
        :return: 行数
        """
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        """
        获取对应行的值
        :param row: 行数
        :return: 对应行的值，列表格式
        """
        row_list = []

        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)

        return row_list

    def excel_write_data(self, row, col, value):
        """
        向 Excel 表格写入数据
        :param row: 行
        :param col: 列
        :param value: 内容
        :return:
        """
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, col, value)
        wb.save(base_path + "/case/case.xlsx")

    def get_column_value(self, key=None):
        """
        获取对应列的值
        :param key: 对应的列，key = "A"
        :return: 对应行列的值，列表格式
        """
        if not key:
            key = "A"

        column_list = []
        column_list_data = self.get_sheet_data()[key]
        for item in column_list_data:
            column_list.append(item.value)
        return column_list

    def get_row_number(self, case_id):
        """
        获取行号
        :param case_id: 用例编号
        :return: 对应行号
        """
        col_data = self.get_column_value()
        if case_id in col_data:
            return col_data.index(case_id) + 1
        return None

    def get_excel_data(self):
        """将 Excel 中每一行的数据存储在一个列表中"""
        data_list = []
        for i in range(2, self.get_rows() + 1):
            data_list.append(self.get_rows_value(i))

        return data_list


excel_data = HandleExcel()
